"""
xl-processor.py

Batch-convert annotation CSVs into Excel (.xlsx) files that already have the
codebook.md columns AND dropdown menus, so a collaborator can just open a file
and start coding — no software beyond Excel / LibreOffice, no code.

What it does:
  - Reads every CSV in INPUT_FOLDER.
  - Writes one .xlsx per CSV into OUTPUT_FOLDER (a NEW folder; the CSVs are never
    modified or overwritten).
  - Appends the codebook annotation columns, each with a dropdown where it applies.
  - Adds a hidden 'lists' sheet (holds the long goal_action list) and a 'codebook'
    sheet (label definitions for quick reference).

You normally only need to set INPUT_FOLDER below.

Requires:  pip install pandas openpyxl
Run:       python xl-processor.py
"""

import os
import glob
import pandas as pd
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment

# ============================ SET THIS ============================
INPUT_FOLDER = "~/Downloads/BBN_Delft_2025/debriefing_sets/validation"        # folder that holds your annotation CSVs
# e.g. a Mac Downloads subfolder:  "~/Downloads/mydata"
# =================================================================
OUTPUT_FOLDER = "~/Downloads/BBN_Delft_2025/debriefing_sets/annotation_xlsx"      # NEW folder for the .xlsx output (CSVs untouched)
FILE_GLOB = "*.csv"                    # which files in INPUT_FOLDER to convert


# --- Codebook (mirrors codebook.md) ---------------------------------------
# code -> (goal title, [action strings])
GOALS = {
    "g1": ("Establish a Supportive and Professional Environment", [
        "Addressed family member by name.",
        "Introduced him/herself by name and role.",
        "Clearly stated the name of the deceased family member.",
        "Sat down. (Body language/Eye contact)",
        "Displayed professional attire/presence.",
        "Handled interruptions in non-disruptive manner.",
        "Conducted interaction in organized manner.",
    ]),
    "g2": ("Assess and Align Expectations", [
        "Ensured all important survivors were present.",
        "Determined knowledge survivors possessed.",
        "Involved me when discussing reason for visit.",
        "Elicited patient perspective of health situation.",
    ]),
    "g3": ("Deliver the News", [
        "Provided appropriate opening statement (warning shot).",
        "Accurately/succinctly chronicled events leading to death.",
        "Used phrase 'dead' or 'died' (avoided euphemisms).",
        "Avoided jargon or explained terms.",
    ]),
    "g4": ("Manage the Emotional Response", [
        "Paused to allow family to assimilate information.",
        "Responded to cues with appropriate touch.",
        "Emotional response did not interfere with communication.",
        "Legitimized my emotions.",
        "Reinforced positive behaviors.",
    ]),
    "g5": ("Ensure Understanding and Facilitate Closure", [
        "Offered viewing of the deceased.",
        "Established availability to answer questions.",
        "Encouraged questions/concerns.",
        "Summarized the interview.",
        "Checked for accuracy during interview.",
        "Reviewed next step(s).",
        "Verified patient's understanding.",
    ]),
}

# goal_action dropdown values: g1-1 ... g5-7, then 'other'
GOAL_ACTIONS = [f"{code}-{i}"
                for code, (_title, acts) in GOALS.items()
                for i in range(1, len(acts) + 1)] + ["other"]

LEVELS = ["R0", "R1", "R2", "R3", "R4"]
VALENCE = ["positive", "negative", "neutral", "mixed"]
YES_NO = ["Yes", "No"]

LEVEL_DEFS = [
    ("R0", "Description — only what happened, or a bare 'I did well', no reason"),
    ("R1", "Reflective Description — what + why (a reason), not questioned"),
    ("R2", "Dialogic — questions own reading / weighs an alternative / reads family's hidden state"),
    ("R3", "Transformative — commits to change or changed view of self ('next time I'll...')"),
    ("R4", "Critical — beyond this encounter: culture / ethics / training system (rare)"),
]
VALENCE_DEFS = [
    ("positive", "appraises the behavior as a strength / good choice"),
    ("negative", "appraises it as a shortcoming / mistake / regret / to change"),
    ("neutral", "pure description, no positive/negative appraisal"),
    ("mixed", "both a positive and a negative in the same reflection"),
]

# annotation columns to append (from codebook.md). value list -> dropdown; None -> free text
ANNOTATION_COLS = [
    ("is_reflection", YES_NO),
    ("goal_action", GOAL_ACTIONS),
    ("level", LEVELS),
    ("valence", VALENCE),
    ("episode", None),          # free text: ep1, ep2, ...
    ("other_summary", None),    # free text
    ("notes", None),            # free text
]


def read_table(path):
    """Read a CSV; fall back to tab-separated if it looks like one column."""
    df = pd.read_csv(path, dtype=str, keep_default_na=False)
    if df.shape[1] == 1:
        alt = pd.read_csv(path, sep="\t", dtype=str, keep_default_na=False)
        if alt.shape[1] > 1:
            df = alt
    return df


def add_codebook_sheet(wb):
    cb = wb.create_sheet("codebook")
    cb.column_dimensions["A"].width = 100
    lines = [
        ("CODEBOOK — quick reference", True),
        ("", False),
        ("is_reflection:  Yes / No", True),
        ("    Yes = speaker recalls / evaluates / reflects on the simulation", False),
        ("    No  = filler, acknowledgement, small talk, logistics", False),
        ("", False),
        ("goal_action:  pick the goal+action code, or 'other'", True),
    ]
    for code, (title, acts) in GOALS.items():
        lines.append((f"{code} — {title}", True))
        for i, act in enumerate(acts, start=1):
            lines.append((f"    {code}-{i}   {act}", False))
    lines += [
        ("other  — fits none of the listed actions (then fill other_summary)", False),
        ("", False),
        ("level  (code the highest that applies):", True),
    ]
    lines += [(f"    {code}  {desc}", False) for code, desc in LEVEL_DEFS]
    lines += [("", False), ("valence:", True)]
    lines += [(f"    {code} — {desc}", False) for code, desc in VALENCE_DEFS]
    lines += [
        ("", False),
        ("episode:  ep1, ep2, ...  same number while the same moment is reflected; new number when it changes", False),
        ("other_summary:  3-7 words, only when goal_action = other", False),
        ("notes:  free text", False),
    ]
    for text, bold in lines:
        cb.append([text])
        if bold:
            cb.cell(row=cb.max_row, column=1).font = Font(bold=True)


def make_xlsx(df, out_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "annotate"

    orig_cols = list(df.columns)
    # skip an annotation column if the source already has that header
    anno_cols = [(name, vals) for name, vals in ANNOTATION_COLS if name not in orig_cols]
    header = orig_cols + [name for name, _ in anno_cols]

    ws.append(header)
    for c in range(1, len(header) + 1):
        ws.cell(row=1, column=c).font = Font(bold=True)

    for _, row in df.iterrows():
        ws.append([row[c] for c in orig_cols] + [""] * len(anno_cols))

    last = max(ws.max_row, 2)   # last data row (guard for empty files)

    # hidden 'lists' sheet for the long goal_action dropdown
    lists = wb.create_sheet("lists")
    lists["A1"] = "goal_action"
    for i, v in enumerate(GOAL_ACTIONS, start=2):
        lists.cell(row=i, column=1, value=v)
    lists.sheet_state = "hidden"
    ga_ref = f"lists!$A$2:$A${len(GOAL_ACTIONS) + 1}"

    # dropdowns on the annotation columns
    start_col = len(orig_cols) + 1
    for offset, (name, values) in enumerate(anno_cols):
        if not values:
            continue
        col = get_column_letter(start_col + offset)
        formula = ga_ref if name == "goal_action" else '"' + ",".join(values) + '"'
        # allow_blank + showErrorMessage=False -> the dropdown SUGGESTS but does not
        # reject, so a multi-goal cell like "g3-2; g4-1" can still be typed in.
        dv = DataValidation(type="list", formula1=formula,
                            allow_blank=True, showErrorMessage=False)
        ws.add_data_validation(dv)
        dv.add(f"{col}2:{col}{last}")

    # formatting: freeze header, widths, wrap the Utterance column
    ws.freeze_panes = "A2"
    anno_names = {name for name, _ in anno_cols}
    utt_idx = None
    for c, name in enumerate(header, start=1):
        letter = get_column_letter(c)
        if name.lower() == "utterance":
            ws.column_dimensions[letter].width = 70
            utt_idx = c
        elif name in anno_names:
            ws.column_dimensions[letter].width = 16
        else:
            ws.column_dimensions[letter].width = 14
    if utt_idx:
        letter = get_column_letter(utt_idx)
        for r in range(2, last + 1):
            ws[f"{letter}{r}"].alignment = Alignment(wrap_text=True, vertical="top")

    add_codebook_sheet(wb)
    wb.save(out_path)


def main():
    in_dir = os.path.expanduser(INPUT_FOLDER)      # expands ~ and ~user
    out_dir = os.path.expanduser(OUTPUT_FOLDER)
    if os.path.abspath(in_dir) == os.path.abspath(out_dir):
        print("INPUT_FOLDER and OUTPUT_FOLDER must differ (do not overwrite the data).")
        return
    os.makedirs(out_dir, exist_ok=True)

    files = sorted(glob.glob(os.path.join(in_dir, FILE_GLOB)))
    if not files:
        print(f"No files matched {os.path.join(in_dir, FILE_GLOB)}")
        return

    print(f"Converting {len(files)} file(s) -> {out_dir}/")
    for path in files:
        try:
            df = read_table(path)
        except Exception as e:
            print(f"  SKIP {os.path.basename(path)} (read error: {e})")
            continue
        base = os.path.splitext(os.path.basename(path))[0]
        out_path = os.path.join(out_dir, base + ".xlsx")
        make_xlsx(df, out_path)
        print(f"  {os.path.basename(path)} -> {os.path.basename(out_path)} ({len(df)} rows)")
    print("Done.")


if __name__ == "__main__":
    main()
